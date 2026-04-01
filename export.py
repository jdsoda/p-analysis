import sys
import json
import os
try:
    from openpyxl import Workbook
except ImportError:
    print("ERROR:openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime

def thin_border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style='medium', color='555555')
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text, bold=True, bg='1F3864', fg='FFFFFF', size=11, center=True):
    cell.value = text
    cell.font  = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.fill  = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center', wrap_text=True
    )
    cell.border = thin_border()

def rupee(n):
    return round(float(n), 2) if n else 0.0

def run_export(json_path, out_path, statement_name="Bank Statement"):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()

    # ─────────────────── SHEET 1: Summary ───────────────────
    ws = wb.active
    ws.title = "Audit Summary"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = "P-ANALYSIS — Bank Statement Audit Report"
    t.font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    t.fill  = PatternFill('solid', start_color='1F3864')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    s = ws['A2']
    s.value = f"Statement: {statement_name}   |   Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    s.font  = Font(name='Arial', size=9, color='555555', italic=True)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Stats row
    total_dr  = sum(rupee(r.get('total_debit',  0)) for r in data)
    total_cr  = sum(rupee(r.get('total_credit', 0)) for r in data)
    total_txn = sum(r.get('count', 0) for r in data)
    flagged   = sum(1 for r in data if r.get('flags'))

    stat_labels = ['Total Parties', 'Total Txns', 'Total Debit (₹)', 'Total Credit (₹)', 'Net (₹)', 'Flagged Parties']
    stat_values = [len(data), total_txn, total_dr, total_cr, round(total_cr - total_dr, 2), flagged]
    stat_bgs    = ['2E4057','2E4057','C0392B','1E8449','1F3864','922B21']

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 22

    ws['A4'].value = "SUMMARY"
    ws['A4'].font  = Font(name='Arial', bold=True, size=8, color='888888')

    for ci, (lbl, val, bg) in enumerate(zip(stat_labels, stat_values, stat_bgs), start=1):
        lc = get_column_letter(ci)
        ws[f'{lc}5'].value = lbl
        ws[f'{lc}5'].font  = Font(name='Arial', bold=True, size=8, color='FFFFFF')
        ws[f'{lc}5'].fill  = PatternFill('solid', start_color=bg)
        ws[f'{lc}5'].alignment = Alignment(horizontal='center', vertical='center')

        ws[f'{lc}6'].value = val
        ws[f'{lc}6'].font  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        ws[f'{lc}6'].fill  = PatternFill('solid', start_color=bg)
        ws[f'{lc}6'].alignment = Alignment(horizontal='center', vertical='center')
        if ci >= 3:
            ws[f'{lc}6'].number_format = '#,##0.00'

    # ── Column headers (row 8) ──
    ws.row_dimensions[8].height = 32
    headers = [
        ('#',           4,  'center'),
        ('Party Name / Narration', 38, 'left'),
        ('Entries',     9,  'center'),
        ('Total Debit (₹)',  16, 'right'),
        ('Total Credit (₹)', 16, 'right'),
        ('Net (₹)',     14, 'right'),
        ('Max Single (₹)',   14, 'right'),
        ('Audit Flags',      35, 'left'),
    ]

    for ci, (hdr, wid, align) in enumerate(headers, start=1):
        lc = get_column_letter(ci)
        apply_header(ws[f'{lc}8'], hdr, center=(align == 'center'))
        ws.column_dimensions[lc].width = wid

    # ── Data rows ──
    FLAG_COLORS = {
        'CHEQUE': ('FDE8E8', 'C0392B'),
        'RETURNED': ('FDE8E8', 'C0392B'),
        'REVERSED': ('FDE8E8', 'C0392B'),
        'HIGH':     ('FFF3CD', 'B7770D'),
        'AGGREGATE':('FFF3CD', 'B7770D'),
        'ROUND':    ('D6EAF8', '1A5276'),
        'FREQUENT': ('D6EAF8', '1A5276'),
    }
    NORMAL_BG = 'FFFFFF'
    ALT_BG    = 'F8F9FA'

    for ri, row in enumerate(data, start=9):
        ws.row_dimensions[ri].height = 18
        flags      = row.get('flags', [])
        flags_text = ' | '.join(flags) if flags else '✓ Normal'

        # Row background
        row_bg = ALT_BG if (ri % 2 == 0) else NORMAL_BG
        flag_bg, flag_fg = NORMAL_BG, '2ECC71'

        for ftext in flags:
            for kw, (fbg, ffg) in FLAG_COLORS.items():
                if kw in ftext.upper():
                    flag_bg, flag_fg = fbg, ffg
                    row_bg = fbg
                    break

        def cell(col, val, fmt=None, bold=False, color='000000', align='left', bg=None):
            lc = get_column_letter(col)
            c = ws[f'{lc}{ri}']
            c.value = val
            c.font  = Font(name='Arial', size=9, bold=bold, color=color)
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.fill  = PatternFill('solid', start_color=(bg or row_bg))
            c.border = thin_border()
            if fmt: c.number_format = fmt

        net = rupee(row.get('net', 0))

        cell(1, ri - 8, align='center', color='888888')
        cell(2, row.get('party','—'), bold=bool(flags), color='1F3864' if not flags else '7B241C')
        cell(3, row.get('count', 0), align='center')
        cell(4, rupee(row.get('total_debit',  0)), fmt='#,##0.00', align='right',
             color='C0392B' if rupee(row.get('total_debit', 0)) > 0 else '888888')
        cell(5, rupee(row.get('total_credit', 0)), fmt='#,##0.00', align='right',
             color='1E8449' if rupee(row.get('total_credit',0)) > 0 else '888888')
        cell(6, net, fmt='#,##0.00', align='right',
             color=('1E8449' if net > 0 else ('C0392B' if net < 0 else '888888')))
        cell(7, rupee(row.get('max_single', 0)), fmt='#,##0.00', align='right')
        cell(8, flags_text, bold=bool(flags),
             color=(flag_fg if flags else '1E8449'), bg=(flag_bg if flags else row_bg))

    # Totals row
    tr = len(data) + 9
    ws.row_dimensions[tr].height = 20
    ws.merge_cells(f'A{tr}:C{tr}')
    tc = ws[f'A{tr}']
    tc.value = f"TOTALS  ({len(data)} parties)"
    tc.font  = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    tc.fill  = PatternFill('solid', start_color='1F3864')
    tc.alignment = Alignment(horizontal='center', vertical='center')

    for col, val in [(4, total_dr), (5, total_cr), (6, round(total_cr - total_dr, 2))]:
        lc = get_column_letter(col)
        c  = ws[f'{lc}{tr}']
        c.value = val
        c.number_format = '#,##0.00'
        c.font  = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill  = PatternFill('solid', start_color='1F3864')
        c.alignment = Alignment(horizontal='right', vertical='center')

    # Freeze panes
    ws.freeze_panes = 'A9'

    # ─────────────────── SHEET 2: Flagged Only ───────────────────
    ws2 = wb.create_sheet("🚩 Flagged Entries")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:H1')
    t2 = ws2['A1']
    t2.value = "FLAGGED AUDIT OBSERVATIONS — Requires CA Attention"
    t2.font  = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    t2.fill  = PatternFill('solid', start_color='922B21')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 28

    for ci, (hdr, wid, align) in enumerate(headers, start=1):
        lc = get_column_letter(ci)
        apply_header(ws2[f'{lc}3'], hdr, bg='7B241C', center=(align == 'center'))
        ws2.column_dimensions[lc].width = wid
    ws2.row_dimensions[3].height = 30

    fi = 4
    flagged_rows = [r for r in data if r.get('flags')]
    for row in flagged_rows:
        ws2.row_dimensions[fi].height = 18
        flags_text = ' | '.join(row['flags'])
        net = rupee(row.get('net', 0))

        def fc2(col, val, fmt=None, bold=False, color='000000', align='left'):
            lc = get_column_letter(col)
            c = ws2[f'{lc}{fi}']
            c.value = val
            c.font  = Font(name='Arial', size=9, bold=bold, color=color)
            c.alignment = Alignment(horizontal=align, vertical='center')
            c.fill  = PatternFill('solid', start_color='FDE8E8')
            c.border = thin_border()
            if fmt: c.number_format = fmt

        fc2(1, fi - 3, align='center', color='888888')
        fc2(2, row.get('party','—'), bold=True, color='922B21')
        fc2(3, row.get('count', 0), align='center')
        fc2(4, rupee(row.get('total_debit',  0)), fmt='#,##0.00', align='right', color='C0392B')
        fc2(5, rupee(row.get('total_credit', 0)), fmt='#,##0.00', align='right', color='1E8449')
        fc2(6, net, fmt='#,##0.00', align='right',
            color=('1E8449' if net > 0 else ('C0392B' if net < 0 else '888888')))
        fc2(7, rupee(row.get('max_single', 0)), fmt='#,##0.00', align='right')
        fc2(8, flags_text, bold=True, color='922B21')
        fi += 1

    if not flagged_rows:
        ws2['A4'].value = "✓ No audit flags detected in this statement."
        ws2['A4'].font  = Font(name='Arial', italic=True, color='1E8449', size=10)

    ws2.freeze_panes = 'A4'

    # ─────────────────── SHEET 3: Legend ───────────────────
    ws3 = wb.create_sheet("Legend & Notes")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 55
    ws3.column_dimensions['C'].width = 30

    apply_header(ws3['A1'], 'FLAG TYPE', bg='1F3864')
    apply_header(ws3['B1'], 'DESCRIPTION', bg='1F3864')
    apply_header(ws3['C1'], 'RELEVANT SECTION', bg='1F3864')

    legend = [
        ('CHEQUE RETURNED / REVERSED', 'Net amount ≈ 0 with multiple entries; possible payment dishonour', 'Sec 138 NI Act'),
        ('HIGH VALUE TXN ≥₹2L',        'Single transaction ≥ ₹2,00,000; must verify cash/mode of payment', 'Sec 269SS / 269T'),
        ('AGGREGATE CREDIT ≥₹10L',     'Total credits from a party exceed ₹10 Lakh; SFT reporting applicable', 'Rule 114E SFT'),
        ('ROUND FIGURE TRANSACTIONS',  '2 or more transactions in exact multiples of ₹1,000 ≥ ₹10,000', 'Audit Scrutiny'),
        ('FREQUENT PARTY',             'More than 10 entries with the same party in the statement', 'Related Party Check'),
    ]
    flag_row_bgs = ['FDE8E8','FFF3CD','FFF3CD','D6EAF8','D6EAF8']
    flag_row_fgs = ['922B21','B7770D','B7770D','1A5276','1A5276']

    for li, (flag, desc, sec) in enumerate(legend, start=2):
        bg, fg = flag_row_bgs[li-2], flag_row_fgs[li-2]
        for col, val in [('A', flag), ('B', desc), ('C', sec)]:
            c = ws3[f'{col}{li}']
            c.value = val
            c.font  = Font(name='Arial', size=9, bold=(col == 'A'), color=fg)
            c.fill  = PatternFill('solid', start_color=bg)
            c.border = thin_border()
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws3.row_dimensions[li].height = 18

    ws3['A9'].value  = "Generated by P-ANALYSIS | For CA / IT Audit use only"
    ws3['A9'].font   = Font(name='Arial', size=8, italic=True, color='AAAAAA')

    wb.save(out_path)
    return out_path


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        json_path = sys.argv[1]
        out_path  = sys.argv[2]
        name      = sys.argv[3] if len(sys.argv) > 3 else "Statement"
        run_export(json_path, out_path, name)
        print("OK:" + out_path)
    else:
        print("ERROR:Missing arguments")
